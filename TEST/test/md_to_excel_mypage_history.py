# -*- coding: utf-8 -*-
"""FO_테스트시나리오_상세_마이페이지_이용내역.md -> .xlsx 변환"""
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다. 설치: pip install openpyxl")
    raise

MD_PATH = Path(__file__).resolve().parent / "FO_테스트시나리오_상세_마이페이지_이용내역.md"
XLSX_PATH = Path(__file__).resolve().parent / "FO_테스트시나리오_상세_마이페이지_이용내역.xlsx"

def parse_table_line(line: str) -> list:
    line = line.strip()
    if not line.startswith("|") or not line.endswith("|"):
        return []
    parts = [c.strip() for c in line[1:-1].split("|")]
    return parts

def cell_text(s: str) -> str:
    """<br> -> 줄바꿈"""
    if not s:
        return s
    return s.replace("<br>", "\n").replace("<br/>", "\n")

def main():
    text = MD_PATH.read_text(encoding="utf-8")
    lines = text.splitlines()
    header = None
    rows = []

    for raw in lines:
        cells = parse_table_line(raw)
        if not cells:
            continue
        if re.match(r"^\-+\s*$", str(cells[0] if cells else "")):
            continue
        if "순번" in (cells[0] if cells else ""):
            header = cells
            continue
        if header and len(cells) >= 1 and cells[0].strip().isdigit():
            # 셀 수가 헤더보다 적을 수 있음(마지막 빈칸 누락 등) → 헤더 개수만큼 채움
            padded = [cell_text(c) for c in cells]
            while len(padded) < len(header):
                padded.append("")
            rows.append(padded[: len(header)])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "25.01. 이용내역"[:31]

    # FO_테스트시나리오_상세_최신.xlsx 와 동일한 형식
    thin = Side(style="thin")
    for c, val in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 22

    # 참고 엑셀과 동일한 열 너비
    col_widths = (8, 8, 9, 8, 19, 11, 8, 8, 8, 8, 8, 8)
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 내용이 보이도록 행 높이 (참고파일은 None; 여기서는 줄 수 반영)
        line_count = max(1, str(row[4] if len(row) > 4 else "").count("\n") + 1)
        ws.row_dimensions[r].height = max(22, min(120, 18 * line_count))

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)
    print(f"저장: {XLSX_PATH}")

if __name__ == "__main__":
    main()

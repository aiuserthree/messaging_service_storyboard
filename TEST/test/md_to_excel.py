# -*- coding: utf-8 -*-
"""FO_테스트시나리오_상세.md -> .xlsx 변환 (섹션별 시트, openpyxl 사용)"""
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("openpyxl이 필요합니다. 설치: pip install openpyxl")
    sys.exit(1)

MD_PATH = Path(__file__).resolve().parent / "FO_테스트시나리오_상세.md"
XLSX_PATH = Path(__file__).resolve().parent / "FO_테스트시나리오_상세.xlsx"

# 엑셀 시트명에 사용 불가 문자
SHEET_FORBIDDEN = re.compile(r'[\\/*?:\[\]]')

def safe_sheet_name(title: str, used: set) -> str:
    """시트명 31자 제한, 불가 문자 제거, 중복 시 번호 붙임"""
    s = SHEET_FORBIDDEN.sub("_", title).strip()[:31]
    if not s:
        s = "시트"
    base, n = s, 0
    while s in used:
        n += 1
        s = (base[: 28] + f"_{n}") if len(base) > 28 else f"{base}_{n}"
        s = s[:31]
    used.add(s)
    return s

def parse_table_line(line: str) -> list:
    """마크다운 테이블 한 줄에서 셀 리스트 추출"""
    line = line.strip()
    if not line.startswith("|") or not line.endswith("|"):
        return []
    parts = [c.strip() for c in line[1:-1].split("|")]
    return parts

def parse_md(path: Path):
    """
    마크다운 전체 파싱.
    returns: [(섹션제목, 헤더리스트, [행1, 행2, ...]), ...]
    """
    text = path.read_text(encoding="utf-8")
    current_section = None
    header = None
    in_table = False
    sections = []  # (section_name, header_list, list of row lists)

    for raw in text.splitlines():
        line = raw.rstrip()

        if re.match(r"^##\s+", line):
            m = re.match(r"^##\s+(.+)", line)
            if m:
                if current_section is not None and header is not None and in_table:
                    pass  # 이미 누적 중이던 섹션은 여기서 저장하지 않고, 새 섹션만 current_section 갱신
                current_section = m.group(1).strip()
                in_table = False
                header = None
            continue

        cells = parse_table_line(line)
        if not cells:
            # 빈 줄이어도 테이블 모드 유지 (다음 줄에 데이터 올 수 있음)
            continue

        if re.match(r"^\-+\s*$", str(cells[0] if cells else "")):
            in_table = True
            continue

        if "순번" in (cells[0] if cells else ""):
            header = cells
            in_table = True
            continue

        if not in_table or not header or not current_section:
            continue

        # 데이터 행: 컬럼 수가 헤더보다 적어도 남는 만큼만 사용
        row = [c.replace("<br>", "\n").replace("<br/>", "\n").strip() if isinstance(c, str) else c for c in cells]
        if len(row) < len(header):
            row += [""] * (len(header) - len(row))
        else:
            row = row[: len(header)]

        # 해당 섹션을 sections에 넣거나 기존에 추가
        found = False
        for i, (sec, h, rows) in enumerate(sections):
            if sec == current_section:
                sections[i][2].append(row)
                found = True
                break
        if not found:
            sections.append([current_section, header, [row]])

    return sections

def main():
    path = MD_PATH
    if not path.exists():
        print(f"파일 없음: {path}")
        sys.exit(1)

    sections = parse_md(path)
    default_header = ["순번", "테스트ID", "테스트케이스명", "페이지/팝업", "단계별 작업 수행 및 확인 내용",
                      "PASS/FAIL", "오류내용", "수정상태", "수정담당자", "처리결과", "최종확인", "비고"]

    wb = openpyxl.Workbook()
    used_names = set()

    for idx, (section_name, header, rows) in enumerate(sections):
        sheet_title = safe_sheet_name(section_name, used_names)
        if idx == 0:
            ws = wb.active
            ws.title = sheet_title
        else:
            ws = wb.create_sheet(title=sheet_title)

        h = header if header else default_header
        for c, val in enumerate(h, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        for r_idx, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c, value=val).alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 70
        for col in "FGHIJKL":
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = 12
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"

    wb.save(XLSX_PATH)
    print(f"저장됨: {XLSX_PATH} (섹션 {len(sections)}개 시트)")


if __name__ == "__main__":
    main()

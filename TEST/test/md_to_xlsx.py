# -*- coding: utf-8 -*-
"""FO_테스트시나리오_상세.md -> FO_테스트시나리오_상세.xlsx 변환 (섹션별 시트)"""
import re
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    exit(1)

MD_PATH = os.path.join(os.path.dirname(__file__), "FO_테스트시나리오_상세.md")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "FO_테스트시나리오_상세.xlsx")

# 시트별 헤더 (섹션 컬럼 없음)
HEADER = ["순번", "테스트ID", "테스트케이스명", "페이지/팝업", "단계별 작업 수행 및 확인 내용",
          "PASS/FAIL", "오류내용", "수정상태", "수정담당자", "처리결과", "최종확인", "비고"]

def sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    """엑셀 시트명에 사용 불가 문자 제거 및 길이 제한"""
    for c in [":", "\\", "/", "?", "*", "[", "]"]:
        name = name.replace(c, " ")
    name = name.strip() or "Sheet"
    return name[:max_len] if len(name) > max_len else name

def parse_row(line: str) -> list:
    """마크다운 테이블 한 행을 셀 리스트로 변환. <br> -> 줄바꿈"""
    line = line.strip()
    if not line.startswith("|") or not line.endswith("|"):
        return None
    cells = [c.strip().replace("<br>", "\n") for c in line.split("|")[1:-1]]
    return cells

def is_sep_line(line: str) -> bool:
    return re.match(r"^\|\s*---+", line.strip()) is not None

def main():
    with open(MD_PATH, "r", encoding="utf-8") as f:
        text = f.read()

    wb = openpyxl.Workbook()
    # 첫 시트는 나중에 제거할 수도 있으므로, 섹션별로 시트 생성 후 첫 시트 이름만 바꿈
    used_sheets = {}
    current_ws = None
    current_section = None
    header_font = Font(bold=True)

    lines = text.splitlines()
    i = 0
    total_rows = 0
    while i < len(lines):
        line = lines[i]
        m = re.match(r"^##\s+(.+)$", line.strip())
        if m:
            current_section = m.group(1).strip()
            sheet_name = sanitize_sheet_name(current_section)
            # 동일 이름 시트가 있으면 숫자 붙임
            if sheet_name in used_sheets:
                used_sheets[sheet_name] += 1
                sheet_name = sanitize_sheet_name(current_section, max_len=28) + f"_{used_sheets[sheet_name]}"
            else:
                used_sheets[sheet_name] = 1
            current_ws = wb.create_sheet(title=sheet_name)
            for col, title in enumerate(HEADER, 1):
                current_ws.cell(row=1, column=col, value=title)
                current_ws.cell(row=1, column=col).font = header_font
                current_ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True, vertical="center")
            current_ws.row_dimensions[1].height = 22
            current_ws.column_dimensions["E"].width = 60  # 단계별 내용
            for c in range(1, len(HEADER) + 1):
                current_ws.column_dimensions[get_column_letter(c)].width = max(8, min(50, len(HEADER[c - 1]) + 2))
            i += 1
            continue

        if line.strip().startswith("| 순번 |") and current_ws is not None:
            i += 1
            if i < len(lines) and is_sep_line(lines[i]):
                i += 1
            while i < len(lines) and (not lines[i].strip() or is_sep_line(lines[i])):
                i += 1
            row_num = 2
            while i < len(lines):
                parsed = parse_row(lines[i])
                if parsed is None:
                    break
                if len(parsed) < 5:
                    i += 1
                    continue
                # parsed: 순번, 테스트ID, 테스트케이스명, 페이지/팝업, 단계별..., PASS/FAIL, 오류내용, 수정상태, 수정담당자, 처리결과, 최종확인, 비고
                out = (parsed + [""] * len(HEADER))[: len(HEADER)]
                for col, val in enumerate(out, 1):
                    current_ws.cell(row=row_num, column=col, value=val)
                    current_ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                row_num += 1
                total_rows += 1
                i += 1
            continue
        i += 1

    # 기본 시트(Sheet) 제거
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(XLSX_PATH)
    print(f"저장: {XLSX_PATH} (시트 {len(wb.sheetnames)}개, 총 {total_rows}행)")

if __name__ == "__main__":
    main()
